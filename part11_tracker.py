"""
part11_tracker.py — automate the Part 11 correlation matrix and portfolio
drawdown log (CLAUDE TRADING SYSTEM v7, Part 11) from real price data.

Replaces manual monthly price entry in
docs/PART11_Correlation_Drawdown_Tracker_v6.xlsx: prices come from the same
yfinance download/cache layer the backtester uses (layer1_data_strategies).

Usage
-----
    # correlation matrix for current holdings (prints flags per Rule 32/34)
    python part11_tracker.py --tickers MFC CRM ADBE ...

    # also fill the Excel tracker's Price History tab (formulas do the rest)
    python part11_tracker.py --tickers MFC CRM ADBE --xlsx

    # drawdown log from a portfolio-value CSV with columns: date,value
    python part11_tracker.py --values portfolio_values.csv

Correlation is pairwise Pearson on the trailing 12 *monthly* returns
(--months to widen), matching the workbook. The rule check flags any
position with correlation > 0.70 against more than one other holding.
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date
from typing import Dict, List, Optional

import numpy as np
import pandas as pd

from v7_scoring import CORRELATION_CAP, correlation_violations, drawdown_response

DOCS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs")
TEMPLATE_XLSX = os.path.join(DOCS_DIR, "PART11_Correlation_Drawdown_Tracker_v6.xlsx")

MAX_TRACKER_TICKERS = 12   # workbook has 12 ticker columns
TRACKER_MONTHS = 25        # workbook holds 25 months -> 24 returns


# --------------------------------------------------------------------------- #
# math (pure, offline — tested in tests/test_v7_scoring.py)
# --------------------------------------------------------------------------- #

def monthly_closes(daily: Dict[str, pd.DataFrame],
                   months: int = TRACKER_MONTHS) -> pd.DataFrame:
    """Month-end closes (last trading day of each month), most recent
    ``months`` rows, one column per ticker."""
    cols = {}
    for t, df in daily.items():
        s = df["Close"].resample("ME").last().dropna()
        cols[t] = s
    out = pd.DataFrame(cols)
    return out.tail(months)


def monthly_returns(closes: pd.DataFrame) -> pd.DataFrame:
    return closes.pct_change().dropna(how="all")


def correlation_matrix(returns: pd.DataFrame, window: int = 12) -> pd.DataFrame:
    """Pairwise Pearson correlation on the trailing ``window`` monthly
    returns (12 by default, per the Rule's '12-month price correlation')."""
    return returns.tail(window).corr().round(2)


def drawdown_log(values: pd.Series) -> pd.DataFrame:
    """Running peak, drawdown %, and Part 11 response band for a series of
    portfolio values indexed by date (mirrors the workbook's log tab)."""
    peak = values.cummax()
    dd = values / peak - 1.0
    rows = []
    for d, v, p, x in zip(values.index, values, peak, dd):
        band = drawdown_response(-x if x < 0 else 0.0)
        rows.append({
            "date": d, "value": v, "peak": p, "drawdown": round(float(x), 4),
            "status": band.status,
            "cash_floor": f">={band.cash_floor_pct:.0f}%"
                          if band.status != "Normal" else "5-15%",
            "new_positions": band.new_positions,
        })
    return pd.DataFrame(rows).set_index("date")


# --------------------------------------------------------------------------- #
# data + xlsx plumbing
# --------------------------------------------------------------------------- #

def fetch_daily(tickers: List[str], refresh: bool = False) -> Dict[str, pd.DataFrame]:
    from layer1_data_strategies import download_data
    # ~3 years of dailies comfortably covers 25 month-ends
    start = (pd.Timestamp.today() - pd.DateOffset(years=3)).strftime("%Y-%m-%d")
    return download_data(tickers, start=start, end="today", min_bars=200,
                         refresh=refresh, verbose=False)


def fill_tracker_xlsx(closes: pd.DataFrame,
                      template: str = TEMPLATE_XLSX,
                      out_path: Optional[str] = None) -> str:
    """Write ticker headers + monthly closes into the Price History tab of a
    copy of the tracker workbook. All downstream tabs are formulas and
    recalculate on open in Excel."""
    import openpyxl

    if len(closes.columns) > MAX_TRACKER_TICKERS:
        raise ValueError(f"tracker holds {MAX_TRACKER_TICKERS} tickers, "
                         f"got {len(closes.columns)}")
    wb = openpyxl.load_workbook(template)
    ws = wb["Price History"]
    # headers: B2.., data: rows 3..27 (25 months), dates in column A
    for j, t in enumerate(closes.columns):
        ws.cell(row=2, column=2 + j, value=t)
    for i, (dt, row) in enumerate(closes.iterrows()):
        r = 3 + i
        ws.cell(row=r, column=1, value=pd.Timestamp(dt).strftime("%Y-%m"))
        for j, t in enumerate(closes.columns):
            v = row[t]
            if pd.notna(v):
                ws.cell(row=r, column=2 + j, value=round(float(v), 2))
    if out_path is None:
        out_path = os.path.join(
            DOCS_DIR, f"PART11_Tracker_filled_{date.today().isoformat()}.xlsx")
    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    ap.add_argument("--tickers", nargs="+", metavar="T",
                    help="current holdings for the correlation matrix")
    ap.add_argument("--months", type=int, default=12,
                    help="correlation window in months (default 12)")
    ap.add_argument("--refresh", action="store_true",
                    help="force fresh price download")
    ap.add_argument("--xlsx", action="store_true",
                    help="also write a filled copy of the Excel tracker")
    ap.add_argument("--values", metavar="CSV",
                    help="portfolio-value CSV (date,value) for the drawdown log")
    args = ap.parse_args(argv)

    if not args.tickers and not args.values:
        ap.error("nothing to do: pass --tickers and/or --values")

    if args.tickers:
        daily = fetch_daily([t.upper() for t in args.tickers],
                            refresh=args.refresh)
        missing = [t for t in (x.upper() for x in args.tickers) if t not in daily]
        if missing:
            print(f"! no usable data for: {', '.join(missing)}", file=sys.stderr)
        closes = monthly_closes(daily)
        rets = monthly_returns(closes)
        corr = correlation_matrix(rets, window=args.months)

        print(f"\nPairwise correlation of monthly returns "
              f"(trailing {args.months} months):\n")
        print(corr.to_string())

        pairs = [(a, b, corr.loc[a, b])
                 for i, a in enumerate(corr.index)
                 for b in corr.columns[i + 1:]
                 if pd.notna(corr.loc[a, b]) and corr.loc[a, b] > CORRELATION_CAP]
        if pairs:
            print(f"\nPairs above the {CORRELATION_CAP:.2f} cap:")
            for a, b, c in sorted(pairs, key=lambda x: -x[2]):
                print(f"  {a} — {b}: {c:.2f}")
        else:
            print(f"\nNo pair exceeds the {CORRELATION_CAP:.2f} cap.")

        viol = correlation_violations(corr)
        if viol:
            print("\nRULE VIOLATIONS (correlated >0.70 with more than one "
                  "other holding — halve the new position or pass):")
            for t, others in viol.items():
                print(f"  {t}: {', '.join(others)}")
        else:
            print("Correlation cap rule: OK "
                  "(no position >0.70 with 2+ other holdings).")

        if args.xlsx:
            path = fill_tracker_xlsx(closes)
            print(f"\nFilled tracker written: {path}")

    if args.values:
        df = pd.read_csv(args.values, parse_dates=[0])
        df.columns = [c.strip().lower() for c in df.columns]
        s = df.set_index(df.columns[0])["value"].astype(float)
        log = drawdown_log(s)
        print("\nPortfolio drawdown log (Part 11 response bands):\n")
        print(log.to_string())
        worst = log["drawdown"].min()
        print(f"\nMax drawdown: {worst:.1%}  |  "
              f"current status: {log['status'].iloc[-1]}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
