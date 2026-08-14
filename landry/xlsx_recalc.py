"""Landry System v1.0 — recalculate and validate a workbook via LibreOffice.

Self-contained and committed to the repo (not borrowed from a session-local
skill install) specifically so it behaves the same on every machine this
codebase runs on. Two lessons are baked into why this exists at all:

1. openpyxl writes formulas as text with no cached value. Anything that
   reads a workbook with data_only=True (pandas, landry.xlsx_io, this
   file's own error scan, Excel's own "quick preview") sees None for
   every formula cell until something actually calculates them.
2. "The XML is well-formed and openpyxl round-trips it" is NOT proof a
   workbook is valid -- Excel's own repair mechanism can still reject a
   file that passes every check available without a real spreadsheet
   engine. Only LibreOffice recalculating it, then Excel actually
   opening it, provides that proof; this module gets you the first half.

The approach (headless LibreOffice, a macro that calls calculateAll()
then store(), invoked via vnd.sun.star.script:) is the same one used by
the project's `xlsx` skill's recalc.py -- reimplemented here rather than
imported because that script requires Python 3.10+ (a keyword argument
this repo's floor of 3.9 doesn't have) and lives outside the repo, in a
session-local skill install whose path isn't the same across machines.

Run: python -m landry.xlsx_recalc <path.xlsx> [timeout_seconds]
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from typing import Optional

MACRO_FILENAME = "Module1.xba"

_MACRO_XML = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def soffice_path() -> Optional[str]:
    return shutil.which("soffice") or shutil.which("libreoffice")


def _stamp(path: str):
    st = os.stat(path)
    return st.st_mtime_ns, st.st_size


def _setup_profile(profile_dir: Path, timeout: int) -> Optional[str]:
    """Bootstrap a throwaway LibreOffice user profile with the recalc
    macro installed. Returns an error string, or None on success."""
    soffice = soffice_path()
    if soffice is None:
        return ("soffice not found on PATH -- run `python -m landry doctor` "
                "for install instructions")
    url = profile_dir.as_uri()
    try:
        subprocess.run(
            [soffice, "--headless", "--terminate_after_init",
             f"-env:UserInstallation={url}"],
            capture_output=True, timeout=timeout)
    except FileNotFoundError:
        return "soffice disappeared from PATH mid-run"
    except subprocess.TimeoutExpired:
        return "LibreOffice timed out creating its profile"

    macro_dir = profile_dir / "user" / "basic" / "Standard"
    if not macro_dir.exists():
        return "LibreOffice did not create a usable profile"
    try:
        (macro_dir / MACRO_FILENAME).write_text(_MACRO_XML)
    except OSError as e:
        return f"could not install the recalculation macro: {e}"
    return None


def recalc(path: str, timeout: int = 30) -> dict:
    """Recalculate every formula in ``path`` in place via headless
    LibreOffice, then scan for Excel error strings. Returns a dict with
    either an "error" key (nothing was recalculated) or "status"
    ("success" | "errors_found"), "total_errors", "total_formulas", and
    "error_summary" ({error_type: [locations]})."""
    if not Path(path).exists():
        return {"error": f"{path} does not exist"}
    abs_path = str(Path(path).absolute())
    if not os.access(abs_path, os.W_OK):
        return {"error": f"{path} is not writable"}

    soffice = soffice_path()
    if soffice is None:
        return {"error": "soffice not found on PATH -- run "
                         "`python -m landry doctor` for install instructions"}

    profile_dir = Path(tempfile.mkdtemp(prefix="landry-recalc-lo-profile-"))
    try:
        started = time.monotonic()
        err = _setup_profile(profile_dir, timeout)
        if err:
            return {"error": err}
        remaining = max(5, int(timeout - (time.monotonic() - started)))

        before = _stamp(abs_path)
        cmd = [soffice, "--headless", "--norestore",
              f"-env:UserInstallation={profile_dir.as_uri()}",
              "vnd.sun.star.script:Standard.Module1.RecalculateAndSave"
              "?language=Basic&location=application", abs_path]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True,
                                    timeout=remaining + 15)
        except subprocess.TimeoutExpired:
            return {"error": f"LibreOffice timed out after {remaining}s; "
                             "nothing was recalculated"}
        except FileNotFoundError:
            return {"error": "soffice not found on PATH"}

        if result.returncode != 0:
            detail = (result.stderr or "").strip() or f"exit {result.returncode}"
            return {"error": f"LibreOffice failed to recalculate: {detail}"}
        if _stamp(abs_path) == before:
            return {"error": "LibreOffice exited cleanly but never rewrote "
                             "the file -- close any other LibreOffice "
                             "instance and retry"}
    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)

    return _scan_errors(path)


def _scan_errors(path: str) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    error_details = {e: [] for e in EXCEL_ERRORS}
    total_errors = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not hasattr(ws, "iter_rows"):
            continue
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    for err in EXCEL_ERRORS:
                        if err in v:
                            error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                            total_errors += 1
                            break
    wb.close()

    wb_f = openpyxl.load_workbook(path, data_only=False)
    formula_count = sum(
        1
        for sheet_name in wb_f.sheetnames
        for row in (wb_f[sheet_name].iter_rows()
                   if hasattr(wb_f[sheet_name], "iter_rows") else [])
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    wb_f.close()

    summary = {e: {"count": len(locs), "locations": locs[:100]}
              for e, locs in error_details.items() if locs}
    return {"status": "success" if total_errors == 0 else "errors_found",
           "total_errors": total_errors, "total_formulas": formula_count,
           "error_summary": summary}


def main() -> int:
    args = sys.argv[1:]
    if not args:
        print("Usage: python -m landry.xlsx_recalc <path.xlsx> [timeout_seconds]")
        return 1
    path = args[0]
    timeout = int(args[1]) if len(args) > 1 else 30
    result = recalc(path, timeout)
    print(json.dumps(result, indent=2))
    return 1 if "error" in result else 0


if __name__ == "__main__":
    sys.exit(main())
