"""Excel round-trip: fill a copy of the Landry workbook from app data, and
seed the score store from a workbook (one-time migration).

Export philosophy: the workbook's formula tabs (Returns, Correlation
Matrix, Scoring computed columns, Action Items) recalculate in Excel on
open — we fill only value cells the Schema Reference tab designates as
inputs: Price History weekly closes, Market Data, the Portfolio Drawdown
Log, and (optionally) analyst scores on the Scoring tab.
"""

from __future__ import annotations

import datetime as _dt
import os
from typing import Dict, List, Mapping, Optional

# Price History layout (Schema Reference): rows 3-162, A=Week Ending,
# B-Q = up to 16 ticker columns, headers in row 2.
PRICE_HISTORY_MAX_TICKERS = 16
PRICE_HISTORY_MAX_ROWS = 160

# Market Data layout: rows 3-27, A-I.
MARKET_COLS = ("price", "volume", "market_cap_m", "pe",
               "wk52_low", "wk52_high", "dividend_yield")

# Scoring tab score columns (same map as xlsx_io reads)
from landry.xlsx_io import _SCORE_COLS  # noqa: E402


def export_workbook(template_path: str,
                    out_path: Optional[str] = None,
                    weekly_closes=None,               # DataFrame, cols=tickers
                    market: Optional[Mapping[str, Mapping]] = None,
                    drawdown=None,                    # DataFrame from regime_frame
                    approved_scores: Optional[Mapping[str, Mapping]] = None,
                    scored_date: Optional[_dt.date] = None) -> str:
    """Fill a copy of the workbook. Only the sections passed are touched.

    approved_scores: {ticker: {indicator: IndicatorScore}} — written to the
    Scoring tab rows whose ticker matches (scores + confidence only; the
    composite/decision/flag columns are Excel formulas and recalculate).
    """
    import openpyxl

    wb = openpyxl.load_workbook(template_path)   # keep formulas intact

    if weekly_closes is not None:
        ws = wb["Price History"]
        cols = list(weekly_closes.columns)[:PRICE_HISTORY_MAX_TICKERS]
        tail = weekly_closes[cols].dropna(how="all").tail(PRICE_HISTORY_MAX_ROWS)
        # clear the data block, then write headers + rows
        # (ws.cell(value=None) is a no-op in openpyxl — assign explicitly)
        for r in range(3, 3 + PRICE_HISTORY_MAX_ROWS):
            for c in range(1, 2 + PRICE_HISTORY_MAX_TICKERS):
                ws.cell(row=r, column=c).value = None
        for j, t in enumerate(cols):
            ws.cell(row=2, column=2 + j, value=t)
        for i, (dt, row) in enumerate(tail.iterrows()):
            r = 3 + i
            ws.cell(row=r, column=1, value=dt.to_pydatetime()
                    if hasattr(dt, "to_pydatetime") else dt)
            for j, t in enumerate(cols):
                v = row[t]
                if v == v and v is not None:      # not NaN
                    ws.cell(row=r, column=2 + j, value=round(float(v), 2))

    if market:
        ws = wb["Market Data"]
        # existing ticker rows: A3..A27
        row_of = {}
        for r in range(3, 28):
            t = ws.cell(row=r, column=1).value
            if t:
                row_of[str(t).strip().upper()] = r
        next_free = max(row_of.values(), default=2) + 1
        for t, m in market.items():
            t = t.upper()
            r = row_of.get(t)
            if r is None and next_free <= 27:
                r, next_free = next_free, next_free + 1
                ws.cell(row=r, column=1, value=t)
            if r is None:
                continue
            mc = m.get("market_cap")
            vals = (m.get("price"), m.get("volume"),
                    (mc / 1e6 if mc else None), m.get("pe"),
                    m.get("wk52_low"), m.get("wk52_high"),
                    m.get("dividend_yield"))
            for j, v in enumerate(vals):
                if v is not None:
                    ws.cell(row=r, column=3 + j, value=v)

    if drawdown is not None and len(drawdown):
        ws = wb["Portfolio Drawdown Log"]
        tail = drawdown.tail(40)
        for i, (dt, row) in enumerate(tail.iterrows()):
            r = 3 + i
            ws.cell(row=r, column=1, value=dt.to_pydatetime()
                    if hasattr(dt, "to_pydatetime") else dt)
            ws.cell(row=r, column=2, value=round(float(row["value"]), 2))
            ws.cell(row=r, column=3, value=round(float(row["peak"]), 2))
            ws.cell(row=r, column=4, value=round(float(row["drawdown"]), 4))
            ws.cell(row=r, column=5, value=str(row["status"]))
            ws.cell(row=r, column=6, value=str(row["cash_floor"]))
            ws.cell(row=r, column=7, value=str(row["new_positions"]))

    if approved_scores:
        ws = wb["Scoring"]
        row_of = {}
        for r in range(3, 28):
            t = ws.cell(row=r, column=1).value
            if t:
                row_of[str(t).strip().upper()] = r
        next_free = max(row_of.values(), default=2) + 1
        for t, scores in approved_scores.items():
            t = t.upper()
            r = row_of.get(t)
            if r is None and next_free <= 27:
                r, next_free = next_free, next_free + 1
                ws.cell(row=r, column=1, value=t)
            if r is None:
                continue
            if scored_date:
                ws.cell(row=r, column=3, value=scored_date)
            for name, col in _SCORE_COLS:
                sc = scores.get(name)
                if sc is not None:
                    ws.cell(row=r, column=col, value=int(sc.score))
                    ws.cell(row=r, column=col + 1, value=sc.confidence)

    if out_path is None:
        stamp = _dt.date.today().isoformat()
        base = os.path.basename(template_path).replace(".xlsx", "")
        out_path = os.path.join(os.path.dirname(template_path),
                                f"{base}_filled_{stamp}.xlsx")
    wb.save(out_path)
    return out_path


def import_scores(workbook_path: str, store, approved_by: str,
                  tickers: Optional[List[str]] = None) -> Dict[str, int]:
    """Seed the ScoreStore from the workbook Scoring tab: every recorded
    indicator score is proposed and immediately approved as source
    "manual" (they ARE the analyst's own numbers — the workbook is the
    original approval). Returns {ticker: indicators imported}."""
    from landry.fundamentals import Draft
    from landry.xlsx_io import read_scoring_tab

    rows = read_scoring_tab(workbook_path)
    wanted = {t.upper() for t in tickers} if tickers else None
    out: Dict[str, int] = {}
    for row in rows:
        if wanted and row.ticker.upper() not in wanted:
            continue
        n = 0
        for name, sc in row.scores.items():
            store.propose(row.ticker, Draft(
                name, sc.score, sc.confidence,
                f"imported from {os.path.basename(workbook_path)} "
                f"(scored {row.date_scored})"), source="manual")
            store.approve(row.ticker, name, approved_by=approved_by)
            n += 1
        out[row.ticker] = n
    return out
