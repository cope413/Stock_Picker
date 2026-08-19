"""Read the tabs of a LANDRY_SYSTEM_WORKBOOK_*.xlsx.

Column maps come from the workbook's own Schema Reference tab. Scoring:
rows 3-45 (43 tickers), cols A-AM. A=Ticker, B=Company, C=Date,
D-M = Tier 1 (5 x score/conf pairs), N=Tier1 WtdAvg, O=Tier1 Contrib,
P-W = Tier 2 (4 x score/conf), X=Tier2 Contrib, Y-AD = Tier 3
(3 x score/conf), AE=Tier3 Contrib, AF=Composite, AG=Decision,
AH-AK = Rule 1-4 flags, AL=Sector, AM=Industry (added 2026-08-18).

Export (``export.py``) fills value cells; ``landry.migrate_to_db`` reads
every tab here to populate ``landry.db`` (see LANDRY_DATABASE_DESIGN.md).
"""

from __future__ import annotations

import glob
import os
from dataclasses import dataclass
from typing import Dict, List, Optional

from landry.scoring import IndicatorScore

WORKBOOK_PREFIX = "LANDRY_SYSTEM_WORKBOOK_"
WORKBOOK_GLOB = f"{WORKBOOK_PREFIX}*.xlsx"


def pick_latest(paths: List[str]) -> Optional[str]:
    """Highest-numbered LANDRY_SYSTEM_WORKBOOK_<N>.xlsx among ``paths``, or None.

    Ignores non-numeric matches (e.g. ..._TEMPLATE_FINAL.xlsx) -- a plain
    lexical sort would rank 'T' above any digit and pick the blank template
    over the live workbook.
    """
    numbered = []
    for path in paths:
        stem = os.path.basename(path)[len(WORKBOOK_PREFIX):-len(".xlsx")]
        if stem.isdigit():
            numbered.append((int(stem), path))
    return max(numbered)[1] if numbered else None


def latest_workbook(repo_dir: str) -> Optional[str]:
    """Highest-numbered LANDRY_SYSTEM_WORKBOOK_<N>.xlsx in repo_dir, or None."""
    return pick_latest(glob.glob(os.path.join(repo_dir, WORKBOOK_GLOB)))


# (indicator, score column index 1-based); confidence is score col + 1
_SCORE_COLS = [
    ("fcf_yield_trend", 4),
    ("revenue_growth_consistency", 6),
    ("competitive_moat", 8),
    ("revenue_visibility", 10),
    ("fcf_margin_trend", 12),
    ("management_quality", 16),
    ("roic_vs_wacc", 18),
    ("relative_strength", 20),
    ("valuation_multiples", 22),
    ("technical_trend", 25),
    ("analyst_consensus", 27),
    ("volume_accumulation", 29),
]

_COL_TIER1_AVG, _COL_COMPOSITE, _COL_DECISION = 14, 32, 33
_COL_RULES = (34, 35, 36, 37)


@dataclass
class WorkbookRow:
    """One scored candidate as recorded in the workbook."""
    ticker: str
    company: str
    date_scored: Optional[object]
    scores: Dict[str, IndicatorScore]          # only the indicators present
    tier1_weighted_average: Optional[float]    # workbook-computed (col N)
    composite: Optional[float]                 # workbook-computed (col AF)
    decision: Optional[str]                    # workbook-computed (col AG)
    rule_flags: tuple                          # workbook cols AH-AK
    sector: Optional[str] = None                # col AL
    industry: Optional[str] = None              # col AM


@dataclass
class Position:
    """One row of the Current Positions tab (equities and cash alike)."""
    account: str
    ticker: str
    description: str
    asset_class: str
    quantity: float
    market_value: float
    pct_of_portfolio: float
    unrealized_pct: Optional[float] = None   # col J, fraction (-0.15 = -15%)
    notes: str = ""


def read_positions(path: str, sheet: str = "Current Positions") -> List[Position]:
    """Current Positions tab: A=Account, B=Ticker, C=Description,
    D=AssetClass, E=Quantity, G=MarketValue, L=%Combined, M=Notes.
    Skips subtotal/total rows (blank ticker) and zero-quantity rows
    (sold positions kept for the record)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    out: List[Position] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or not r[0] or not r[1]:
            continue
        qty = float(r[4] or 0)
        if qty <= 0:
            continue
        out.append(Position(
            account=str(r[0]).strip(), ticker=str(r[1]).strip(),
            description=str(r[2] or "").strip(),
            asset_class=str(r[3] or "").strip(), quantity=qty,
            market_value=float(r[6] or 0),
            pct_of_portfolio=float(r[11] or 0),
            unrealized_pct=float(r[9]) if r[9] is not None else None,
            notes=str(r[12] or "").strip(),
        ))
    wb.close()
    return out


def equity_weights(positions: List[Position]) -> dict:
    """{ticker: combined portfolio fraction} for Equity rows, summed
    across accounts (NVDA appears in both)."""
    out: dict = {}
    for p in positions:
        if p.asset_class == "Equity":
            out[p.ticker] = out.get(p.ticker, 0.0) + p.pct_of_portfolio
    return out


def total_portfolio_value(positions: List[Position]) -> float:
    """Sum of market_value across every position and account (Part 7's
    drawdown series is portfolio-wide, cash included)."""
    return sum(p.market_value for p in positions)


def read_drawdown_log(path: str,
                      sheet: str = "Portfolio Drawdown Log") -> "pd.Series":
    """Date -> Portfolio Value from the tab's own manually-entered history
    (cols A, B; data from row 3). Empty Series if nothing's been logged
    yet. Only the raw value column is read -- Running Peak/Drawdown/Status
    are workbook formulas derived from it, not a second source of truth."""
    import openpyxl
    import pandas as pd

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    dates, values = [], []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] is None or r[1] is None:
            continue
        dates.append(r[0])
        values.append(float(r[1]))
    wb.close()
    return pd.Series(values, index=pd.DatetimeIndex(dates),
                     name="value").sort_index()


def read_scoring_tab(path: str, sheet: str = "Scoring") -> List[WorkbookRow]:
    import openpyxl  # local import: optional dependency for non-Excel use

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows: List[WorkbookRow] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or not r[0]:
            continue
        scores: Dict[str, IndicatorScore] = {}
        for name, col in _SCORE_COLS:
            raw, conf = r[col - 1], r[col]
            if raw is None:
                continue
            scores[name] = IndicatorScore(int(raw), str(conf or "M"))
        rows.append(WorkbookRow(
            ticker=str(r[0]).strip(),
            company=str(r[1] or "").strip(),
            date_scored=r[2],
            scores=scores,
            tier1_weighted_average=(float(r[_COL_TIER1_AVG - 1])
                                    if r[_COL_TIER1_AVG - 1] is not None else None),
            composite=(float(r[_COL_COMPOSITE - 1])
                       if r[_COL_COMPOSITE - 1] is not None else None),
            decision=(str(r[_COL_DECISION - 1]).strip()
                      if r[_COL_DECISION - 1] else None),
            rule_flags=tuple(str(r[c - 1] or "OK").strip() for c in _COL_RULES),
            sector=(str(r[37]).strip() if len(r) > 37 and r[37] else None),
            industry=(str(r[38]).strip() if len(r) > 38 and r[38] else None),
        ))
    wb.close()
    return rows


# --------------------------------------------------------------------------- #
# Remaining tabs -- readers for landry.migrate_to_db (see
# LANDRY_DATABASE_DESIGN.md). Plain dicts, not dataclasses: these feed one
# consumer (the migration script) rather than being a shared contract like
# Position/WorkbookRow above.
# --------------------------------------------------------------------------- #

def _open(path: str, sheet: str):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return wb, wb[sheet]


def read_market_data(path: str, sheet: str = "Market Data") -> List[dict]:
    """A=Ticker,B=Company,C=Price,D=Volume,E=MarketCap($M),F=P/E,
    G=52wkLow,H=52wkHigh,I=DivYield. Rows 3-27 (Schema Reference); bounded
    at row 27 because explanatory footnote text starts right below the
    table and would otherwise be read as a bogus data row. Manual-entry
    tab, often empty (data_auto.py fetches this live instead)."""
    wb, ws = _open(path, sheet)
    out = []
    for r in ws.iter_rows(min_row=3, max_row=27, values_only=True):
        if not r or not r[0]:
            continue
        out.append(dict(
            ticker=str(r[0]).strip(), company=str(r[1] or "").strip(),
            price=r[2], volume=r[3], market_cap_m=r[4], pe=r[5],
            wk52_low=r[6], wk52_high=r[7], dividend_yield=r[8]))
    wb.close()
    return out


def read_price_history(path: str, sheet: str = "Price History") -> List[dict]:
    """Row 2 = ticker headers (col A = 'Week Ending'), data from row 3.
    Long/tidy output: one row per (ticker, week_ending)."""
    wb, ws = _open(path, sheet)
    header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    tickers = [str(h).strip() if h else None for h in header]
    out = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] is None:
            continue
        week_ending = r[0]
        for i in range(1, len(r)):
            if i >= len(tickers) or not tickers[i] or r[i] is None:
                continue
            out.append(dict(ticker=tickers[i], week_ending=week_ending,
                            close=float(r[i])))
    wb.close()
    return out


def read_monitor_notes(path: str,
                       sheet: str = "Monitor & Recheck Triggers") -> List[dict]:
    """Only the genuinely-manual fields; LastComposite/CurrentPrice/
    DaysSinceScore etc. are derived from composite_history + market_data at
    read time, not stored twice. A=Ticker,B=Category,K=InsiderY/N,
    L=InsiderNote,M=AnalystShift,O=RecheckStatus,P=Notes."""
    wb, ws = _open(path, sheet)
    out = []
    for r in ws.iter_rows(min_row=3, max_row=42, values_only=True):
        if not r or not r[0]:
            continue
        out.append(dict(
            ticker=str(r[0]).strip(), category=str(r[1] or "").strip(),
            insider_flag=r[10], insider_note=r[11], analyst_shift_flag=r[12],
            recheck_status=r[14], notes=r[15]))
    wb.close()
    return out


def read_watchlist(path: str, sheet: str = "Watch List Tracker") -> List[dict]:
    """A=Ticker,B=Company,C=Status,D=EntryDate,E=EntryScore,
    H=RemediationPlanY/N,I=90DayDeadline,K=ActionStatus,L=Notes.
    CurrentScore/CurrentScoreDate/DaysRemaining omitted -- derived."""
    wb, ws = _open(path, sheet)
    out = []
    for r in ws.iter_rows(min_row=3, max_row=22, values_only=True):
        if not r or not r[0]:
            continue
        out.append(dict(
            ticker=str(r[0]).strip(), status=r[2], entry_date=r[3],
            entry_score=r[4], remediation_plan_yn=r[7],
            deadline_90day=r[8], action_status=r[10], notes=r[11]))
    wb.close()
    return out


def read_performance_tracking(path: str,
                              sheet: str = "Performance Tracking") -> List[dict]:
    """A=Ticker,B=Company,C=EntryDate,D=EntryPrice,E=EntryScore,
    F=EntryConfidence,G=EntryBand,H=SPYatEntry,I=Status,J=ExitDate,
    K=ExitPrice,L=ExitReason. M-Q (current/exit price, SPY, returns)
    omitted -- computed at report time from current market data. Bounded
    to the PerformanceTrackingTable's own range (row 33): explanatory
    footnote text starts right below it and would otherwise be read as a
    bogus data row."""
    wb, ws = _open(path, sheet)
    out = []
    for r in ws.iter_rows(min_row=3, max_row=33, values_only=True):
        if not r or not r[0]:
            continue
        out.append(dict(
            ticker=str(r[0]).strip(), entry_date=r[2], entry_price=r[3],
            entry_score=r[4], entry_confidence=r[5], entry_band=r[6],
            spy_at_entry=r[7], status=r[8], exit_date=r[9],
            exit_price=r[10], exit_reason=r[11]))
    wb.close()
    return out


def read_holding_monitor(path: str, sheet: str = "Holding Monitor") -> List[dict]:
    """Rows 4+, A-AB. A=Ticker,B=Company,C=PositionPct,D=MaxFullPct.
    E-S = 5 fundamental indicators x [Prior,Current,Flag] (FCF Yield, Moat,
    ROIC-WACC spread, Mgmt Quality, Rev Growth) -> nested 'indicators' list.
    T=Debt/FCF,U=P/FCF,V=FCFGrowth,W=ImpliedReturn,X=CurrentTier,
    Y=PriorTier,Z=ValuationFlags,AA=HoldThroughY/N,AB=ActionStatus."""
    wb, ws = _open(path, sheet)
    indicator_names = ("fcf_yield", "moat", "roic_wacc_spread",
                       "mgmt_quality", "rev_growth")
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not r[0]:
            continue
        indicators = []
        for i, name in enumerate(indicator_names):
            base = 4 + i * 3
            indicators.append(dict(indicator_name=name, prior_value=r[base],
                                   current_value=r[base + 1], flag=r[base + 2]))
        out.append(dict(
            ticker=str(r[0]).strip(), company=str(r[1] or "").strip(),
            position_pct=r[2], max_full_pct=r[3], indicators=indicators,
            debt_fcf=r[19], p_fcf=r[20], fcf_growth=r[21],
            implied_return=r[22], current_tier=r[23], prior_tier=r[24],
            valuation_flags=r[25], hold_through_yn=r[26], action_status=r[27]))
    wb.close()
    return out


def read_implied_return_scenarios(
        path: str, sheet: str = "Implied-Return Calculator") -> List[dict]:
    """Rows 4+, A-U, 2-row merged header. A=Ticker,B=Company,C=Price.
    D-H=Base[FCF5,Mult,Dist,Return,Tag]. I-M=Bear[same]. N-R=Bull[same].
    One output row per (ticker, scenario)."""
    wb, ws = _open(path, sheet)
    scenarios = (("base", 3), ("bear", 8), ("bull", 13))
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not r[0]:
            continue
        ticker, company = str(r[0]).strip(), str(r[1] or "").strip()
        for name, base in scenarios:
            if all(r[base + k] is None for k in range(5)):
                continue
            out.append(dict(
                ticker=ticker, company=company, scenario=name,
                fcf_yr5=r[base], terminal_mult=r[base + 1],
                distributions=r[base + 2], implied_return=r[base + 3],
                tag=r[base + 4]))
    wb.close()
    return out


def read_entry_checklist(path: str, sheet: str = "Entry Checklist") -> List[dict]:
    """A=Ticker,B=Company,C-R = Rules 5-13 results (see Schema Reference for
    the exact per-column rule mapping)."""
    wb, ws = _open(path, sheet)
    out = []
    for r in ws.iter_rows(min_row=3, max_row=45, values_only=True):
        if not r or not r[0]:
            continue
        out.append(dict(
            ticker=str(r[0]).strip(), company=str(r[1] or "").strip(),
            rule5_composite=r[2], rule6_tier1=r[3], rule7_no_tier1_eq1=r[4],
            rule8_200wk=r[5], rule9_macd=r[6], staging_result=r[7],
            rule10_binary_risk=r[8], risk_in_thesis_yn=r[9],
            rule10_result=r[10], rule11_bear_case=r[11],
            rule12_scenario_tags=r[12], p_fcf_current=r[13],
            consensus_fcf_growth_2yr=r[14], rule13_valuation_ceiling=r[15],
            entry_authorized=r[16], recommended_action=r[17]))
    wb.close()
    return out


def read_journal(path: str, sheet: str = "Journal") -> List[dict]:
    """A=Date,B=Ticker,C=Notes. Freeform append-only log."""
    wb, ws = _open(path, sheet)
    out = []
    for r in ws.iter_rows(min_row=3, max_row=302, values_only=True):
        if not r or r[0] is None:
            continue
        out.append(dict(date=r[0], ticker=(str(r[1]).strip() if r[1] else None),
                        notes=r[2]))
    wb.close()
    return out


def read_tax_loss_carryforward(path: str,
                               sheet: str = "Current Positions") -> List[dict]:
    """Loose cells outside the table: O2/P2 = Short-term label/value,
    O3/P3 = Long-term label/value."""
    wb, ws = _open(path, sheet)
    out = []
    for term, row in (("short", 2), ("long", 3)):
        val = ws.cell(row=row, column=16).value   # col P
        if val is not None:
            out.append(dict(term=term, amount=val))
    wb.close()
    return out


def read_drawdown_log_full(path: str,
                           sheet: str = "Portfolio Drawdown Log") -> List[dict]:
    """Full column set for DB migration (read_drawdown_log above stays the
    stable, minimal contract for Part 7 regime logic and isn't changed
    here). A=Date,B=Value,C=RunningPeak,D=Drawdown%,E=Status,
    F=RequiredCashFloor,G=NewPositionInitiation,H=Notes."""
    wb, ws = _open(path, sheet)
    out = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] is None or r[1] is None:
            continue
        out.append(dict(
            date=r[0], portfolio_value=float(r[1]),
            running_peak=r[2], drawdown_pct=r[3], status=r[4],
            cash_floor=r[5], new_position_rule=r[6], notes=r[7]))
    wb.close()
    return out


def read_positions_full(path: str, sheet: str = "Current Positions") -> List[dict]:
    """Full column set for DB migration (read_positions above stays a
    stable, minimal contract for existing callers -- correlation checks,
    sizing -- and isn't changed here). A=Account,B=Ticker,C=Description,
    D=AssetClass,E=Quantity,F=Price,G=MarketValue,H=CostBasis,
    I=UnrealizedGL$,J=UnrealizedGL%,K=%ofAccount,L=%ofCombined,
    M=Classification."""
    wb, ws = _open(path, sheet)
    out = []
    for r in ws.iter_rows(min_row=3, max_row=39, values_only=True):
        if not r or not r[0] or not r[1]:
            continue
        qty = float(r[4] or 0)
        if qty <= 0:
            continue
        out.append(dict(
            account=str(r[0]).strip(), ticker=str(r[1]).strip(),
            description=str(r[2] or "").strip(),
            asset_class=str(r[3] or "").strip(), quantity=qty,
            price=r[5], market_value=r[6], cost_basis=r[7],
            unrealized_gl=r[8], unrealized_gl_pct=r[9],
            pct_of_account=r[10], pct_of_combined=r[11],
            classification=str(r[12] or "").strip()))
    wb.close()
    return out
